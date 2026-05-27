"""
formato/ica_formato.py
----------------------
Aplicacion de colores y estilos a las hojas del Excel de ICA (NADF-009).
"""

import json
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# Carga de colores desde config.json
# ----------------------------------------------------------------------------
# Al importar este modulo, se lee la configuracion para obtener los colores
# asociados a cada rango de ICA. La clave en el JSON tiene formato "0-50",
# se convierte a tupla de enteros (0, 50) para facilitar la comparacion.
with open("config.json", "r", encoding="utf-8") as _f:
    _cfg = json.load(_f)

COLORES_NADF: dict = {
    tuple(map(int, k.split("-"))): v for k, v in _cfg["NADF"]["colores"].items()
}


def obtener_color_ica(valor: int) -> str | None:
    """
    Devuelve el codigo hexadecimal de color NADF-009 para un valor ICA dado.

    Parametros
    ----------
    valor : int
        Valor del ICA (0 a 500)

    Retorna
    -------
    str or None
        Codigo hexadecimal sin '#', ej. "9ACA3C" para el rango 0-50.
        None si el valor no cae en ningun rango definido.
    """
    for (lo, hi), color in COLORES_NADF.items():
        if lo <= valor <= hi:
            return color
    return None


def aplicar_formato_ica(ws) -> None:
    """
    Aplica formato visual a una hoja de Excel que contiene datos ICA.

    Acciones realizadas:
      - Alineacion centrada con ajuste de texto para todas las celdas
      - Encabezado (primera fila) en negrita
      - Fondo de color segun rango ICA en las celdas de datos (numericas)
      - Ancho de columna automatico basado en el contenido (maximo 50)
      - Alto de fila fijo de 25 puntos

    Parametros
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Hoja de trabajo a la que aplicar el formato.
    """
    # ------------------------------------------------------------------------
    # 1. Alineacion global: centrado horizontal y vertical con ajuste de texto
    # ------------------------------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

    # ------------------------------------------------------------------------
    # 2. Encabezado (fila 1) en negrita
    # ------------------------------------------------------------------------
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ------------------------------------------------------------------------
    # 3. Colores de fondo segun el valor ICA
    #    Se itera desde la fila 2 (datos) hasta el final.
    #    Se aplica color solo a celdas numericas (ICA) que no sean NaN.
    #    La columna 1 es la fecha/hora; se salta.
    # ------------------------------------------------------------------------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if (
                cell.column > 1
                and isinstance(cell.value, (int, float))
                and not pd.isna(cell.value)
            ):
                color = obtener_color_ica(int(cell.value))
                if color:
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    # ------------------------------------------------------------------------
    # 4. Ajustar ancho de columnas
    #    Calcula la longitud maxima del contenido en cada columna (incluyendo
    #    el encabezado) y establece el ancho a ese valor + 4, con un tope de 50.
    # ------------------------------------------------------------------------
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 50
        )

    # ------------------------------------------------------------------------
    # 5. Alto de fila fijo de 25 puntos
    # ------------------------------------------------------------------------
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25
