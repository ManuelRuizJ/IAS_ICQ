"""
almacenamiento/exportador.py
-----------------------------
Prepara y escribe los DataFrames finales en Excel.

Nomenclatura de columnas (nueva)
---------------------------------
  AIRE_<unidad>_<contaminante>_<estacion>      → categoría NOM-172
  CANTIDAD_<unidad>_<contaminante>_<estacion>  → concentración
  ICA_<contaminante>_<estacion>                → valor ICA NADF-009

Ejemplo: AIRE_ppb_O3_AGUA SANTA, CANTIDAD_ppm_CO_ATLIXCO, ICA_PM10_BINE
"""

import re
import pandas as pd
from openpyxl import load_workbook

from procesadores.nom import peor_categoria
from formato.ica_formato import aplicar_formato_ica
from formato.aire_formato import aplicar_formato_aire


# ── Patrones de columna ──────────────────────────────────────────────────────
# AIRE_<unidad>_<contaminante>_<estacion>
_PAT_AIRE = re.compile(r"^AIRE_([^_]+)_([^_]+)_(.+)$")
# CANTIDAD_<unidad>_<contaminante>_<estacion>
_PAT_CANT = re.compile(r"^CANTIDAD_([^_]+)_([^_]+)_(.+)$")
# ICA_<contaminante>_<estacion>
_PAT_ICA = re.compile(r"^ICA_([^_]+)_(.+)$")


# ── Ordenación de columnas ───────────────────────────────────────────────────


def ordenar_columnas_ica(df: pd.DataFrame) -> list:
    """Columnas ICA ordenadas por (estación, contaminante)."""
    pares = set()
    for col in df.columns:
        m = _PAT_ICA.match(col)
        if m:
            pares.add((m.group(1), m.group(2)))  # (contaminante, estacion)

    pares_ord = sorted(pares, key=lambda x: (x[1], x[0]))
    cols_ord = [f"ICA_{cont}_{est}" for cont, est in pares_ord]
    resto = [c for c in df.columns if c not in cols_ord]
    return cols_ord + resto


def ordenar_columnas_aire(df: pd.DataFrame) -> list:
    """
    Columnas AIRE/DIARIO intercaladas: AIRE_<u>_X_Y, CANTIDAD_<u>_X_Y, …
    ordenadas por (estación, contaminante), con 'Calidad del aire' al final.
    """
    pares = set()
    for col in df.columns:
        m = _PAT_AIRE.match(col)
        if m:
            # (unidad, contaminante, estacion)
            pares.add((m.group(1), m.group(2), m.group(3)))

    pares_ord = sorted(pares, key=lambda x: (x[2], x[1]))  # por estacion, cont
    cols_ord = []
    for unidad, cont, est in pares_ord:
        cat = f"AIRE_{unidad}_{cont}_{est}"
        cant = f"CANTIDAD_{unidad}_{cont}_{est}"
        if cat in df.columns:
            cols_ord.append(cat)
        if cant in df.columns:
            cols_ord.append(cant)

    if "Calidad del aire" in df.columns:
        cols_ord.append("Calidad del aire")

    resto = [c for c in df.columns if c not in cols_ord]
    return cols_ord + resto


# ── Extracción de hojas por estación ────────────────────────────────────────


def extraer_estaciones_ica(df: pd.DataFrame) -> dict:
    """Divide el DataFrame ICA general en sub-DataFrames por estación."""
    estaciones = set()
    for col in df.columns:
        m = _PAT_ICA.match(col)
        if m:
            estaciones.add(m.group(2))

    resultado = {}
    for est in sorted(estaciones):
        cols = [
            c
            for c in df.columns
            if _PAT_ICA.match(c) and _PAT_ICA.match(c).group(2) == est
        ]
        resultado[est] = df[cols].copy()
    return resultado


def extraer_estaciones_aire(
    df: pd.DataFrame, orden_cat: dict, suficiencia: float
) -> dict:
    """
    Divide el DataFrame AIRE/DIARIO en sub-DataFrames por estación,
    recalculando 'Calidad del aire' para cada una.
    """
    estaciones = set()
    for col in df.columns:
        m = _PAT_AIRE.match(col) or _PAT_CANT.match(col)
        if m:
            estaciones.add(m.group(3))  # grupo 3 = estacion

    resultado = {}
    for est in sorted(estaciones):
        cols_cat = [
            c
            for c in df.columns
            if _PAT_AIRE.match(c) and _PAT_AIRE.match(c).group(3) == est
        ]
        cols_cant = [
            c
            for c in df.columns
            if _PAT_CANT.match(c) and _PAT_CANT.match(c).group(3) == est
        ]

        df_est = df[cols_cat + cols_cant].copy()

        if cols_cat:
            df_est["Calidad del aire"] = peor_categoria(
                [df_est[c] for c in cols_cat],
                orden_cat,
                umbral=suficiencia,
            )

        # Reordenar columnas de la estación
        pares = []
        for col in cols_cat:
            m = _PAT_AIRE.match(col)
            if m:
                pares.append((m.group(1), m.group(2)))  # (unidad, contaminante)

        cols_ord = []
        for unidad, cont in sorted(pares, key=lambda x: x[1]):
            cat = f"AIRE_{unidad}_{cont}_{est}"
            cant = f"CANTIDAD_{unidad}_{cont}_{est}"
            if cat in df_est.columns:
                cols_ord.append(cat)
            if cant in df_est.columns:
                cols_ord.append(cant)
        if "Calidad del aire" in df_est.columns:
            cols_ord.append("Calidad del aire")

        resultado[est] = df_est[cols_ord]

    return resultado


# ── Escritura del Excel final ────────────────────────────────────────────────


def guardar_diccionario_excel(
    archivo: str,
    diccionario_dfs: dict,
    tipo: str,
    nombre_indice: str,
) -> None:
    """
    Escribe todas las hojas del diccionario en el archivo Excel y aplica
    el formato de color/estilo.

    tipo : 'ICA', 'AIRE' o 'DIARIO'
    """
    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        for nombre_hoja, df in diccionario_dfs.items():
            nombre_hoja = nombre_hoja[:31]

            if tipo == "DIARIO":
                df_export = df.reset_index()
                df_export.rename(columns={"index": nombre_indice}, inplace=True)
                df_export[nombre_indice] = pd.to_datetime(
                    df_export[nombre_indice]
                ).dt.strftime("%Y-%m-%d")
                df_export.to_excel(writer, sheet_name=nombre_hoja, index=False)
            else:
                df.index.name = nombre_indice
                df.to_excel(writer, sheet_name=nombre_hoja, index=True)

    wb = load_workbook(archivo)
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        if tipo == "ICA":
            aplicar_formato_ica(ws)
        else:
            aplicar_formato_aire(ws)
    wb.save(archivo)
