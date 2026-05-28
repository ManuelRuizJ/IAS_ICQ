"""
formato/aire_formato.py
-----------------------
Estilos para hojas AIRE Y SALUD y DIARIO (NOM-172-2023).

Reconoce la nueva nomenclatura:
  AIRE_<unidad>_<contaminante>_<estacion>     → columna de categoria
  CANTIDAD_<unidad>_<contaminante>_<estacion> → columna numerica
  Calidad del aire                            → columna de categoria global
"""

import json
import re
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# Carga de colores desde config.json
# ----------------------------------------------------------------------------
with open("config.json", "r", encoding="utf-8") as _f:
    _cfg = json.load(_f)


# Mapeo categoria -> color hexadecimal (sin '#')
# Ejemplo: "Buena": "00E400", "Aceptable": "FFFF00", etc.
COLORES_NOM: dict = _cfg["NOM"]["colores"]


# Patron para extraer el contaminante de una columna CANTIDAD_*
# Formato: CANTIDAD_<unidad>_<contaminante>_<estacion>
# Este patron captura el contaminante (grupo 1)
_PAT_CANT = re.compile(r"^CANTIDAD_[^_]+_([^_]+)_")


def aplicar_formato_aire(ws) -> None:
    """
    Aplica colores NOM-172 y formato numerico a una hoja AIRE o DIARIO.

    Acciones:
      1. Identifica las columnas de categoria (AIRE_*, Calidad del aire, etc.)
      2. Alineacion centrada con ajuste de texto
      3. Encabezado en negrita
      4. Fondo de color segun la categoria en las celdas de categoria
      5. Formato numerico en columnas CANTIDAD_* (0, 0.00 o 0.000 segun contaminante)
      6. Ancho de columna automatico
      7. Alto de fila fijo de 25

    Parametros
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Hoja de trabajo a formatear.
    """
    # ------------------------------------------------------------------------
    # 1. Identificar columnas de categoria
    #    Son aquellas cuyo nombre empieza con "AIRE_" (y no contienen "CANTIDAD")
    #    o son exactamente "Calidad del aire" o "Calidad del aire zona".
    # ------------------------------------------------------------------------
    columnas_cat = []
    for col in ws.iter_cols(min_row=1, max_row=1):
        nombre = col[0].value
        if isinstance(nombre, str):
            if (nombre.startswith("AIRE_") and "CANTIDAD" not in nombre) or nombre in (
                "Calidad del aire",
                "Calidad del aire zona",
            ):
                columnas_cat.append(col[0].column)

    # ------------------------------------------------------------------------
    # 2. Alineacion global: centrado horizontal y vertical con ajuste de texto
    # ------------------------------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

    # ------------------------------------------------------------------------
    # 3. Encabezado (primera fila) en negrita
    # ------------------------------------------------------------------------
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ------------------------------------------------------------------------
    # 4. Colores de fondo para celdas de categoria
    #    Para cada fila a partir de la 2, si la celda esta en una columna
    #    de categoria y su valor (categoria) existe en COLORES_NOM, se
    #    aplica fondo solido con ese color.
    #    El color de fuente es negro para Buena/Aceptable, blanco para las demas.
    # ------------------------------------------------------------------------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in columnas_cat and cell.value in COLORES_NOM:
                cell.fill = PatternFill(
                    start_color=COLORES_NOM[cell.value],
                    end_color=COLORES_NOM[cell.value],
                    fill_type="solid",
                )
                # Contraste: texto negro para fondos claros, blanco para oscuros
                color_fuente = (
                    "000000" if cell.value in ("Buena", "Aceptable") else "FFFFFF"
                )
                cell.font = Font(bold=True, color=color_fuente)

    # ------------------------------------------------------------------------
    # 5. Formato numerico en columnas CANTIDAD_*
    #    Determina el numero de decimales segun el contaminante:
    #      O3, NO2, SO2 → 3 decimales
    #      CO           → 2 decimales
    #      PM10, PM2.5  → 0 decimales
    # ------------------------------------------------------------------------
    for col in ws.columns:
        nombre = col[0].value
        if isinstance(nombre, str) and nombre.startswith("CANTIDAD_"):
            # Extraer el contaminante usando la expresion regular
            m = _PAT_CANT.match(nombre)
            cont = m.group(1) if m else ""
            if cont in ("O3", "NO2", "SO2"):
                fmt = "0.000"
            elif cont == "CO":
                fmt = "0.00"
            # Aplicar formato a todas las celdas de datos (desde fila 2)
            else:
                fmt = "0"
            for cell in col[1:]:
                if cell.value is not None:
                    cell.number_format = fmt

    # ------------------------------------------------------------------------
    # 6. Ajustar ancho de columnas automaticamente
    # ------------------------------------------------------------------------
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 50
        )

    # ------------------------------------------------------------------------
    # 7. Alto de fila fijo de 25 puntos
    # ------------------------------------------------------------------------
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25
