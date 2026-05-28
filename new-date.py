"""
new-date.py
-----------
Script auxiliar para cambiar el año de las fechas en un archivo Excel
de datos de calidad del aire.

Problema original: El software exportador puede escribir las filas
de medianoche como "DD/MM/YYYY 24:00" con el año incorrecto
(ej. 2026 cuando los datos son de 2024).

Este script permite corregir el año de todas las fechas de un archivo
Excel existente y guardar una copia con el año ajustado.

Uso:
    1. Configurar archivo_original, archivo_nuevo y nuevo_anio
    2. Ejecutar: python new-date.py

Nota: Solo modifica la primera columna (columna A) que contiene fecha y hora.
"""

import openpyxl
from datetime import datetime

# ----------------------------------------------------------------------------
# CONFIGURACION
# ----------------------------------------------------------------------------
archivo_original = "datos/datos_calidad_aire_2025.xlsx"  # Archivo fuente
archivo_nuevo = "datos/datos_calidad_aire_2026.xlsx"  # Archivo de salida
nuevo_anio = 2026  # Año al que se cambiaran las fechas

# ----------------------------------------------------------------------------
# PROCESAMIENTO
# ----------------------------------------------------------------------------

# Cargar el libro de trabajo original
wb = openpyxl.load_workbook(archivo_original)
ws = wb.active  # Trabajar con la hoja activa (primera hoja)

# Las filas 1-3 contienen encabezados (estaciones, contaminantes, unidades).
# Los datos reales comienzan en la fila 4 (primera fila de fecha/hora).
for row in range(4, ws.max_row + 1):
    celda = ws.cell(row, 1)  # Primera columna = "Fecha & Hora"
    if celda.value and isinstance(celda.value, str):
        try:
            # Parsear fecha en formato "dd/mm/yyyy HH:MM"
            dt = datetime.strptime(celda.value, "%d/%m/%Y %H:%M")
            # Reemplazar solo el año
            dt = dt.replace(year=nuevo_anio)
            # Volver a escribir en el mismo formato de string
            celda.value = dt.strftime("%d/%m/%Y %H:%M")
        except ValueError:
            # Si la celda no contiene una fecha valida (ej. texto de estadisticas),
            # se omite sin error.
            pass

# Guardar el archivo modificado
wb.save(archivo_nuevo)
print(f"Archivo generado: {archivo_nuevo} con fechas del año {nuevo_anio}")
