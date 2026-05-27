"""
almacenamiento/exportador.py
-----------------------------
Prepara y escribe los DataFrames finales en Excel.
Nueva nomenclatura:
  AIRE_<contaminante>_<estacion>                → categoria
  CANTIDAD_<unidad>_<contaminante>_<estacion>   → concentracion
  ICA_<contaminante>_<estacion>                 → valor ICA
"""

import re
import pandas as pd
from openpyxl import load_workbook

from procesadores.nom import peor_categoria
from formato.ica_formato import aplicar_formato_ica
from formato.aire_formato import aplicar_formato_aire

# ----------------------------------------------------------------------------
# REGEX patrones de columnas para identificar cada tipo
# ----------------------------------------------------------------------------
# AIRE_<contaminante>_<estacion>
_PAT_AIRE = re.compile(r"^AIRE_([^_]+)_(.+)$")
# CANTIDAD_<unidad>_<contaminante>_<estacion>
_PAT_CANT = re.compile(r"^CANTIDAD_([^_]+)_([^_]+)_(.+)$")
# ICA_<contaminante>_<estacion>
_PAT_ICA = re.compile(r"^ICA_([^_]+)_(.+)$")


def ordenar_columnas_ica(df: pd.DataFrame) -> list:
    """
    Ordena las columnas ICA por (estacion, contaminante).

    Parametros
    ----------
    df : pd.DataFrame
        DataFrame con columnas ICA_*

    Retorna
    -------
    list
        Lista de nombres de columnas ordenadas: primero todas las ICA_*
        ordenadas por estacion y dentro por contaminante, luego el resto.
    """
    pares = set()
    for col in df.columns:
        m = _PAT_ICA.match(col)
        if m:
            pares.add((m.group(1), m.group(2)))             # (contaminante, estacion)
    # Ordenar por estacion (segundo elemento) y luego por contaminante
    pares_ord = sorted(pares, key=lambda x: (x[1], x[0]))
    cols_ord = [f"ICA_{cont}_{est}" for cont, est in pares_ord]
    resto = [c for c in df.columns if c not in cols_ord]
    return cols_ord + resto


def ordenar_columnas_aire(df: pd.DataFrame) -> list:
    """
    Ordena las columnas AIRE/CANTIDAD por (estacion, contaminante),
    intercalando categoria y cantidad para cada contaminante.

    Parametros
    ----------
    df : pd.DataFrame
        DataFrame con columnas AIRE_* y CANTIDAD_*

    Retorna
    -------
    list
        Lista de nombres de columnas ordenadas: para cada par (cont, est)
        primero la columna AIRE_* y luego su correspondiente CANTIDAD_*,
        despues "Calidad del aire" si existe, y finalmente el resto.
    """
    # Extraer todos los pares (contaminante, estacion) de las columnas AIRE_
    pares = set()
    for col in df.columns:
        m = _PAT_AIRE.match(col)
        if m:
            pares.add((m.group(1), m.group(2)))
    pares_ord = sorted(pares, key=lambda x: (x[1], x[0]))
    cols_ord = []
    for cont, est in pares_ord:
        col_cat = f"AIRE_{cont}_{est}"
        # Buscar la columna de cantidad correspondiente (termina con _{cont}_{est})
        col_cant_candidates = [
            c
            for c in df.columns
            if c.endswith(f"_{cont}_{est}") and c.startswith("CANTIDAD_")
        ]
        col_cant = col_cant_candidates[0] if col_cant_candidates else None
        cols_ord.append(col_cat)
        if col_cant:
            cols_ord.append(col_cant)
    if "Calidad del aire" in df.columns:
        cols_ord.append("Calidad del aire")
    resto = [c for c in df.columns if c not in cols_ord]
    return cols_ord + resto


def extraer_estaciones_ica(df: pd.DataFrame) -> dict:
    """
    Extrae sub-DataFrames por estacion para los datos ICA.

    Parametros
    ----------
    df : pd.DataFrame
        DataFrame general con columnas ICA_*

    Retorna
    -------
    dict
        Diccionario {nombre_estacion: DataFrame con solo las columnas ICA de esa estacion}
    """
    estaciones = set()
    for col in df.columns:
        m = _PAT_ICA.match(col)
        if m:
            estaciones.add(m.group(2))      # grupo 2 es el nombre de estacion
    resultado = {}
    for est in sorted(estaciones):
        cols = [c for c in df.columns if c.endswith(f"_{est}") and c.startswith("ICA_")]
        resultado[est] = df[cols].copy()
    return resultado


def extraer_estaciones_aire(
    df: pd.DataFrame, orden_cat: dict, suficiencia: float
) -> dict:
    """
    Divide el DataFrame AIRE/DIARIO en sub-DataFrames por estacion,
    recalculando 'Calidad del aire' para cada una.

    Parametros
    ----------
    df : pd.DataFrame
        DataFrame general con columnas AIRE_* y CANTIDAD_*
    orden_cat : dict
        Mapeo categoria -> valor numerico para calcular peor categoria
    suficiencia : float
        Fraccion minima de columnas que deben tener dato para calcular
        la categoria global de la estacion.

    Retorna
    -------
    dict
        Diccionario {nombre_estacion: DataFrame con las columnas de esa estacion
        mas la columna "Calidad del aire" recalculada solo para esa estacion}
    """
    estaciones = set()
    # Buscar estaciones en columnas AIRE_*
    estaciones = set()
    for col in df.columns:
        m = _PAT_AIRE.match(col)
        if m:
            estaciones.add(m.group(2))
        else:
            m = _PAT_CANT.match(col)
            if m:
                estaciones.add(m.group(3))      # grupo 3 es la estacion
    resultado = {}
    for est in sorted(estaciones):
        cols_cat = [
            c
            for c in df.columns
            if _PAT_AIRE.match(c) and _PAT_AIRE.match(c).group(2) == est
        ]
        cols_cant = [
            c
            for c in df.columns
            if _PAT_CANT.match(c) and _PAT_CANT.match(c).group(3) == est
        ]
        df_est = df[cols_cat + cols_cant].copy()
        if cols_cat:
            df_est["Calidad del aire"] = peor_categoria(
                [df_est[c] for c in cols_cat], orden_cat, umbral=suficiencia
            )

        # Reordenar columnas de la estacion (intercaladas categoria y cantidad
        pares = []
        for col in cols_cat:
            m = _PAT_AIRE.match(col)
            if m:
                pares.append((m.group(1), m.group(2)))      # (contaminante, estacion)
        cols_ord = []
        for cont, est2 in sorted(pares, key=lambda x: x[0]):
            col_cat = f"AIRE_{cont}_{est2}"
            col_cant = [
                c
                for c in df_est.columns
                if c.endswith(f"_{cont}_{est2}") and c.startswith("CANTIDAD_")
            ]
            col_cant = col_cant[0] if col_cant else None
            cols_ord.append(col_cat)
            if col_cant:
                cols_ord.append(col_cant)
        if "Calidad del aire" in df_est.columns:
            cols_ord.append("Calidad del aire")
        resultado[est] = df_est[cols_ord]
    return resultado


def guardar_diccionario_excel(
    archivo: str,
    diccionario_dfs: dict,
    tipo: str,
    nombre_indice: str,
) -> None:
    """
    Escribe todas las hojas del diccionario en el archivo Excel y aplica
    el formato de color/estilo.

    Parametros
    ----------
    archivo : str
        Ruta del archivo Excel de salida
    diccionario_dfs : dict
        Mapeo nombre_hoja -> DataFrame
    tipo : str
        'ICA', 'AIRE' o 'DIARIO' (determina que formato aplicar)
    nombre_indice : str
        Nombre que se le da a la columna de fecha/hora en el Excel
        ('Fecha & Hora' para horarios, 'Fecha' para diarios)
    """
    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        for nombre_hoja, df in diccionario_dfs.items():
            # Limitar nombre de hoja a 31 caracteres (maximo de Excel)
            nombre_hoja = nombre_hoja[:31]
            if tipo == "DIARIO":
                # Para reportes diarios: la fecha esta en el indice, pero se quiere
                # guardar como columna para mejor legibilidad.
                df_export = df.reset_index()
                df_export.rename(columns={"index": nombre_indice}, inplace=True)
                 # Formatear fecha como YYYY-MM-DD sin hora
                df_export[nombre_indice] = pd.to_datetime(
                    df_export[nombre_indice]
                ).dt.strftime("%Y-%m-%d")
                df_export.to_excel(writer, sheet_name=nombre_hoja, index=False)
            # Para datos horarios: mantener la fecha como indice
            else:
                df.index.name = nombre_indice
                df.to_excel(writer, sheet_name=nombre_hoja, index=True)

    # Una vez guardado, abrir el libro con openpyxl para aplicar formatos
    wb = load_workbook(archivo)
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        if tipo == "ICA":
            aplicar_formato_ica(ws)
        else:
            aplicar_formato_aire(ws)
    wb.save(archivo)
